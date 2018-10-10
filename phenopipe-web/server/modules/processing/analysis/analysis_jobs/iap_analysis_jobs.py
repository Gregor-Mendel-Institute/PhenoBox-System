import os
from datetime import datetime

import grpc
from openpyxl import Workbook
from rq import get_current_job

from server.gen import phenopipe_pb2_grpc, phenopipe_pb2, phenopipe_iap_pb2_grpc, phenopipe_iap_pb2
from server.models import AnalysisModel, ImageModel, TimestampModel
from server.modules.processing.analysis.analysis_jobs.job_type import JobType
from server.modules.processing.analysis.analysis_jobs.worker_extensions import get_session, get_log_store, \
    get_redis_connection, get_grpc_channel
from server.modules.processing.analysis.analysis_task import AnalysisTask
from server.utils.util import get_local_path_from_smb

_iap_meta_cols = [
    'import file',
    'plant ID',
    'replicate',
    'time',
    'time unit',
    'imaging time',
    'measurement tool',
    'camera.position',
    'rotation degree',
    'species',
    'genotype',
    'variety',
    'growth conditions',
    'treatment'
]


# TODO use a class for this?
# TODO Add proper application logging
def create_return_object(type, timestamp_id, response):
    """
       Convenience method to create a proper return object for the RQ Job function

       :param type: The type of the Job. Instance of :class:`~server.modules.postprocessing.postprocessing_jobs.job_type.JobType`
       :param timestamp_id: The ID of the processed :class:`~server.models.timestamp_model.TimestampModel` instance
       :param response: A dict containing return values that are job specific

       :return: A dict containing the given values with the keys: 'type', 'timestamp_id', 'response'
       """
    ret = dict()
    ret['type'] = type
    ret['timestamp_id'] = timestamp_id
    ret['response'] = response
    return ret


def create_iap_import_sheet(timestamp_id, path):
    """
    Creates the metatdata sheet 'fileimport.xlsx' for IAP at the specified path

    :param timestamp_id: The ID of the :class:`~server.models.timestamp_model.TimestampModel` instance for which the file should be generated
    :param path: The destination path to which the 'fileimport.xlsx' should be saved

    :return:None
    """
    session = get_session()
    timestamp = session.query(TimestampModel).get(timestamp_id)
    if timestamp.iap_exp_id is not None:
        return False
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'meta-data'
    for row in sheet.iter_rows(min_row=1, max_col=len(_iap_meta_cols), max_row=1):
        for index, cell in enumerate(row):
            cell.value = _iap_meta_cols[index]

    start_date = timestamp.experiment.start_of_experimentation
    if start_date is None:
        start_date = timestamp.experiment.start_date

    row_index = 2
    for snap in timestamp.snapshots:
        images = snap.images.all()
        for index, row in enumerate(
                sheet.iter_rows(min_row=row_index, max_col=len(_iap_meta_cols),
                                max_row=row_index + len(images) - 1)):
            # Assumption based on iap_exp_id check: there are only raw images and no segmented ones
            image = images[index]
            row[0].value = image.filename  # filename
            row[1].value = snap.plant.full_name  # plant id
            row[2].value = snap.id  # replicate id has to be unique for each plant
            time_diff = image.created_at - start_date
            row[3].value = time_diff.days  # time
            row[4].value = 'day'  # time unit
            row[5].value = image.created_at  # imaging time
            row[6].value = snap.measurement_tool
            row[7].value = snap.camera_position
            row[8].value = image.angle
            group = snap.plant.sample_group
            row[9].value = group.species
            row[10].value = group.genotype
            row[11].value = group.variety
            row[12].value = group.growth_conditions
            row[13].value = group.treatment
        row_index += len(images)

    wb.save(os.path.join(path, 'fileimport.xlsx'))


# @job('analysis', connection=redis_db)
def invoke_iap_import(timestamp_id, experiment_name, coordinator, scientist, local_path, path, username,
                      task_key):
    """
    This Methods represents an RQ Job workload. It should be enqueued into the RQ Analysis Queue and processed by an according worker

    Handles the invokation of data import into IAP on the IAP server and fetches the result information afterwards.
    The received information is then entered into the database accordingly

    :param timestamp_id: The ID of the :class:`~server.models.timestamp_model.TimestampModel` instance which should be imported
    :param experiment_name: The name of the experiment to import
    :param coordinator: The name of the experiment coordinator
    :param scientist: The name of the scientist carrying out the experiment
    :param local_path: The path to the data on the local system
    :param path: The SMB url representing the location of the data
    :param username: The username of the user invoking this job
    :param task_key: The redis key of the :class:`~server.modules.analysis.analysis_task.AnalysisTask` to which this job belongs

    :return: A dict containing the 'experiment_id' (nested in the 'response' key) returned by IAP
    """
    print('EXECUTE IMPORT')
    job = get_current_job()
    log_store = get_log_store()
    task = AnalysisTask.from_key(get_redis_connection(), task_key)
    channel = get_grpc_channel()
    iap_stub = phenopipe_iap_pb2_grpc.PhenopipeIapStub(channel)
    pipe_stub = phenopipe_pb2_grpc.PhenopipeStub(channel)
    log_store.put(job.id, 'Started Import Job', 0)
    task.update_message('Started Import Job')
    log_store.put(job.id, 'Create Metadata File')
    task.update_message('Create Metadata File')
    create_iap_import_sheet(timestamp_id, local_path)
    log_store.put(job.id, 'Metadata File Created')
    task.update_message('Metadata File Created')
    try:
        log_store.put(job.id, 'Import data into IAP')
        task.update_message('Import data into IAP')
        response = iap_stub.ImportExperiment(
            phenopipe_iap_pb2.ImportRequest(path=path, experiment_name=experiment_name,
                                            coordinator_name=coordinator,
                                            user_name=scientist)
        )

        remote_job_id = response.job_id
        print(remote_job_id)
        request = phenopipe_pb2.WatchJobRequest(
            job_id=remote_job_id
        )
        status = pipe_stub.WatchJob(request)

        for msg in status:
            print(msg.message.decode('string-escape'))
            log_store.put(job.id, msg.message.decode('string-escape'), msg.progress)

        response = iap_stub.FetchImportResult(
            phenopipe_pb2.FetchJobResultRequest(job_id=remote_job_id)
        )
        session = get_session()
        timestamp = session.query(TimestampModel).get(timestamp_id)
        timestamp.iap_exp_id = response.experiment_id
        session.commit()
        log_store.put(job.id, 'Finished Import Job', 100)
        task.update_message('Finished Import Job')
        return create_return_object(JobType.iap_import, timestamp_id, {'experiment_id': response.experiment_id})
    except grpc.RpcError as e:
        if e.code() == grpc.StatusCode.ALREADY_EXISTS:
            session = get_session()
            timestamp = session.query(TimestampModel).get(timestamp_id)
            timestamp.iap_exp_id = e.initial_metadata()[0][1]
            session.commit()
            return create_return_object(JobType.iap_import, timestamp_id, {'experiment_id': timestamp.iap_exp_id})
        else:
            task.update_message('Import Job Failed')
            log_store.put(job.id, e.details(), 0)
            print(e.details())
            raise


# @job('analysis', connection=redis_db)
def invoke_iap_analysis(analysis_id, timestamp_id, username, task_key,
                        experiment_id=None):
    """
    This Methods represents an RQ Job workload. It should be enqueued into the RQ Analysis Queue and processed by an according worker

    Handles the invocation of data analysis in IAP on the IAP server and fetches the result information afterwards.
    The received information is then entered into the database accordingly

    :param analysis_id: The ID of the :class:`~server.models.analysis_model.AnalysisModel`
    :param timestamp_id: The ID of the :class:`~server.models.timestamp_model.TimestampModel` instance which should be analyzed
    :param username: The username of the user invoking this job
    :param analysis_status_id: The ID of the :class:`~server.utils.redis_status_cache.status_object.StatusObject` to which this job belongs
    :param experiment_id: The IAP ID of this experiment. If this is None the job will assume that the job it depended on
        has returned the experiment id in its response object with the key 'experiment_id'

    :return: A dict containing the 'result_id' from IAP, the used 'pipeline_id', 'started_at' and 'finished_at' timestamps.
        (All nested inside the 'response' key)
    """
    print('EXECUTE ANALYSIS')
    job = get_current_job()
    log_store = get_log_store()
    task = AnalysisTask.from_key(get_redis_connection(), task_key)
    channel = get_grpc_channel()
    iap_stub = phenopipe_iap_pb2_grpc.PhenopipeIapStub(channel)
    pipe_stub = phenopipe_pb2_grpc.PhenopipeStub(channel)
    if experiment_id is None:
        experiment_iap_id = job.dependency.result['response'][
            'experiment_id']  # TODO rename experiment_id to experiment_iap_id
    else:
        experiment_iap_id = experiment_id
    log_store.put(job.id, 'Started Analysis Job', 0)
    task.update_message('Started Analysis Job')
    session = get_session()
    # TODO Consider DB errors
    analysis = session.query(AnalysisModel).get(analysis_id)
    started_at = datetime.utcnow()
    analysis.started_at = started_at
    session.commit()
    try:
        response = iap_stub.AnalyzeExperiment(
            phenopipe_iap_pb2.AnalyzeRequest(experiment_id=experiment_iap_id, pipeline_id=analysis.pipeline_id)
        )
        remote_job_id = response.job_id
        request = phenopipe_pb2.WatchJobRequest(
            job_id=remote_job_id
        )
        status = pipe_stub.WatchJob(request)
        for msg in status:
            print(msg.message.decode('string-escape'))
            log_store.put(job.id, msg.message.decode('string-escape'), msg.progress)

        response = iap_stub.FetchAnalyzeResult(
            phenopipe_pb2.FetchJobResultRequest(job_id=remote_job_id)
        )
        finished_at = datetime.utcnow()

        analysis.iap_id = response.result_id
        analysis.finished_at = finished_at
        session.commit()
        log_store.put(job.id, 'Finished Analysis Job', 100)
        task.update_message('Finished Analysis Job')
        return create_return_object(JobType.iap_analysis, timestamp_id,
                                    {'result_id': response.result_id, 'started_at': started_at,
                                     'finished_at': finished_at, 'pipeline_id': analysis.pipeline_id})
    except grpc.RpcError as e:
        session.delete(session.query(AnalysisModel).get(analysis.id))
        session.commit()
        log_store.put(job.id, e.details(), 0)
        task.update_message('Analysis Job Failed')
        print(e.details())
        raise


# @job('analysis', connection=redis_db)
def invoke_iap_export(timestamp_id, output_path, username, shared_folder_map, task_key, analysis_iap_id=None):
    """
    This Methods represents an RQ Job workload. It should be enqueued into the RQ Analysis Queue and processed by an according worker

    Handles the invokation of data export of an IAP analysis on the IAP server and fetches the result information afterwards.
    The received information is then entered into the database accordingly

    :param timestamp_id: The ID of the :class:`~server.models.timestamp_model.TimestampModel` instance to which the data belongs
    :param output_path: The path, as SMB URL, where the data should be exported to
    :param username: The username of the user invoking this job
    :param analysis_status_id: The ID of the :class:`~server.utils.redis_status_cache.status_object.StatusObject` to which this job belongs
    :param shared_folder_map: A dict containing a mapping between SMB URLs and local paths representing the corresponding mount points
    :param analysis_iap_id: The IAP ID of the analysis on the IAP server

    :return: a dict containing the 'analysis_id' for which the data has been exported
        and the 'path' to which the results have been exported. (All nested inside the 'response' key)
    """
    print('EXECUTE EXPORT')
    job = get_current_job()
    log_store = get_log_store()
    task = AnalysisTask.from_key(get_redis_connection(), task_key)
    channel = get_grpc_channel()
    iap_stub = phenopipe_iap_pb2_grpc.PhenopipeIapStub(channel)
    pipe_stub = phenopipe_pb2_grpc.PhenopipeStub(channel)

    if analysis_iap_id is None:
        analysis_iap_id = job.dependency.result['response']['result_id']
    else:
        analysis_iap_id = analysis_iap_id
    log_store.put(job.id, 'Started Export Job', 0)
    task.update_message('Started Export Job')
    try:
        response = iap_stub.ExportExperiment(
            phenopipe_iap_pb2.ExportRequest(experiment_id=analysis_iap_id, destination_path=output_path)
        )
        remote_job_id = response.job_id
        request = phenopipe_pb2.WatchJobRequest(
            job_id=remote_job_id
        )
        status = pipe_stub.WatchJob(request)
        for msg in status:
            print(msg.message.decode('string-escape'))
            log_store.put(job.id, msg.message.decode('string-escape'), msg.progress)

        response = iap_stub.FetchExportResult(
            phenopipe_pb2.FetchJobResultRequest(job_id=remote_job_id)
        )
        session = get_session()
        analysis = session.query(AnalysisModel) \
            .filter(AnalysisModel.timestamp_id == timestamp_id) \
            .filter(AnalysisModel.iap_id == analysis_iap_id) \
            .one()

        log_store.put(job.id, 'Received Results. Started to parse and add information', 90)
        task.update_message('Received Results. Started to parse and add information')
        image_path = get_local_path_from_smb(response.image_path, shared_folder_map)
        print('Image Path: {}'.format(image_path))
        # TODO handle DB errors
        for image_name in os.listdir(image_path):
            print('Image Name: {}'.format(image_name))
            # Extract information from filename
            snapshot_id, _, new_filename = image_name.partition('_')
            _, _, angle = os.path.splitext(image_name)[0].rpartition('_')

            img = ImageModel(snapshot_id, response.image_path, new_filename, angle, 'segmented')
            session.add(img)
            # rename file and remove the snapshot id
            os.rename(os.path.join(image_path, image_name), os.path.join(image_path, new_filename))
        analysis.export_path = response.path
        analysis.exported_at = datetime.utcnow()
        session.commit()
        log_store.put(job.id, 'Finished Export Job', 100)
        task.update_message('Finished Export Job')
        return create_return_object(JobType.iap_export, timestamp_id,
                                    {'analysis_id': analysis.id, 'path': response.path})
    except grpc.RpcError as e:
        log_store.put(job.id, e.details(), 0)
        task.update_message('Export Job Failed')
        print(e.details())
        raise


def dummy_job(name, task_key):
    print('EXECUTE DUMMY')
    job = get_current_job()
    log_store = get_log_store()
    task = AnalysisTask.from_key(get_redis_connection(), task_key)
    log_store.put(job.id, 'Started Dummy Job ({})'.format(name), 0)
    task.update_message('Started Dummy Job ({})'.format(name))
    status = [('{} {}'.format('Status', str(i)), i) for i in range(0, 10)]
    for msg, progress in status:
        log_store.put(job.id, msg, progress)
    log_store.put(job.id, 'Finished Dummy Job ({})'.format(name), 100)
    task.update_message('Finished Dummy Job ({})'.format(name))
    return 'whuiii'
